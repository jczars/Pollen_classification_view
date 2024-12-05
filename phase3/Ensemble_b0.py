#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Nov 4 10:08:02 2024

@author: jczars
"""

import os, sys
import tensorflow as tf
#from tensorflow.keras.preprocessing.image import ImageDataGenerator
from keras.preprocessing.image import ImageDataGenerator
import yaml
import pandas as pd
from openpyxl import Workbook
import argparse
os.environ["tf_gpu_allocator"]="cuda_malloc_async"


from models import voto_majoritary_041124 as vote
from models import sound_test_finalizado, models_train
#from models import reports_build as reports

def list_model_dirs(base_model_dir: str) -> list:
    """
    Lista todos os diretórios de modelos dentro do diretório base.

    Parameters
    ----------
    base_model_dir : str
        Diretório base onde os diretórios dos modelos estão localizados.

    Returns
    -------
    list
        Lista com os nomes dos diretórios dos modelos.
    """
    try:
        model_dirs = [d for d in os.listdir(base_model_dir) if os.path.isdir(os.path.join(base_model_dir, d))]
        return model_dirs
    except FileNotFoundError:
        print(f"Erro: Diretório base '{base_model_dir}' não encontrado.")
        return []
    except Exception as e:
        print(f"Ocorreu um erro ao listar os diretórios dos modelos: {e}")
        return []

def load_model(model_path: str, verbose: int = 0) -> tf.keras.Model | None:
    """
    Carrega um modelo específico a partir do caminho fornecido.
    """
    try:
        print(f"Verificando modelo: {model_path}")
        
        if 'vit' in model_path.lower():
            if verbose > 0:
                print(f"[INFO] Carregando modelo ViT de {model_path}")
            model = models_train.load_model_vit(model_path, verbose=verbose)  # Não passa safe_mode
        else:
            print(f"[INFO] Carregando modelo padrão de {model_path}")
            model = models_train.load_model(model_path, verbose=verbose)  # Para modelos padrão, o safe_mode é útil

        if verbose > 0:
            print(f"Modelo carregado de {model_path}")
        return model
    except Exception as e:
        print(f"Erro ao carregar o modelo de '{model_path}': {e}")
        return None


def load_data_test(data_base_dir: str, K: int, BATCH: int, INPUT_SIZE: tuple, verbose: int = 0) -> tf.keras.utils.Sequence | None:
    """
    Carrega os dados de teste a partir do diretório especificado com base no parâmetro k-fold.

    Parameters
    ----------
    data_base_dir : str
        Diretório base onde as pastas de dados estão localizadas.
    K : int
        Número k-fold.
    BATCH : int
        Tamanho do lote.
    INPUT_SIZE : tuple
        Dimensões de entrada, altura e largura.
    verbose : int, optional
        Modo de verbosidade. Se > 0, imprime mensagens de progresso.

    Returns
    -------
    test_generator : DirectoryIterator ou None
        Gerador de dados de teste, se bem-sucedido; None caso contrário.
    """
    test_dir = os.path.join(data_base_dir, f'Test/k{K}')
    
    try:
        if verbose > 0:
            print(f'Carregando dados de teste do diretório: {test_dir}')

        idg = ImageDataGenerator(rescale=1. / 255)
        test_generator = idg.flow_from_directory(
            directory=test_dir,
            target_size=INPUT_SIZE,
            color_mode="rgb",
            batch_size=BATCH,
            class_mode="categorical",
            shuffle=False,
            seed=42)
        return test_generator
    except Exception as e:
        print(f"Erro ao carregar os dados de teste de '{test_dir}': {e}")
        return None

def load_config(config_path: str, verbose: int = 0) -> dict | None:
    """
    Carrega a configuração de um arquivo YAML.

    Parameters
    ----------
    config_path : str
        Caminho para o arquivo de configuração YAML.
    verbose : int, optional
        Modo de verbosidade. Se > 0, imprime mensagens de progresso.

    Returns
    -------
    dict ou None
        Dicionário de configuração carregado se bem-sucedido; None caso contrário.
    """
    try:
        with open(config_path, 'r') as file:
            config = yaml.safe_load(file)
        if verbose > 0:
            print(f'Configuração carregada de {config_path}')
        return config
    except FileNotFoundError:
        print(f"Erro: Arquivo de configuração '{config_path}' não encontrado.")
        return None
    except yaml.YAMLError as e:
        print(f"Erro YAML ao carregar a configuração: {e}")
        return None
    except Exception as e:
        print(f"Ocorreu um erro ao carregar a configuração: {e}")
        return None

def create_excel_if_not_exists(file_path: str):
    """
    Cria um arquivo Excel vazio caso ele não exista.

    Parameters
    ----------
    file_path : str
        Caminho para o arquivo Excel.
    """
    if not os.path.exists(file_path):
        wb = Workbook()
        wb.save(file_path)
        print(f"Arquivo Excel criado em: '{file_path}'")



def predict_and_save(model, test_data, k, model_name, file_path="test_results.xlsx", sheet_name=None):
    """
    Predicts and saves results to an Excel file.

    Parameters
    ----------
    model : tf.keras.Model
        The model used for prediction.
    test_data : DirectoryIterator
        Test data generator.
    k : int
        k-fold number.
    model_name : str
        Model directory name (used as the sheet name).
    file_path : str, optional
        Path to the Excel file for saving results.
    sheet_name : str, optional
        Name of the sheet where results will be saved.
    """
    results = []
    predictions = model.predict(test_data)
    y_true = test_data.classes
    class_labels = list(test_data.class_indices.keys())
    filenames = test_data.filenames

    for idx, pred in enumerate(predictions):
        y_pred = pred.argmax()
        prob = pred[y_pred]
        label_true = y_true[idx]
        label_pred = y_pred
        sit = "C" if label_true == label_pred else "E"
        
        result = {
            "k": k,
            "Index": idx,
            "Filename": filenames[idx],
            "y_true": class_labels[label_true],
            "y_pred": class_labels[label_pred],
            "Probability": prob,
            "Sit": sit
        }
        results.append(result)

    df = pd.DataFrame(results)

    # Salva os resultados no Excel
    if not os.path.exists(file_path):
        # Cria o arquivo se ele não existir
        with pd.ExcelWriter(file_path, mode='w', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=model_name, index=False)
    else:
        # Adiciona dados à planilha existente
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            if model_name in writer.book.sheetnames:
                # Adiciona abaixo dos dados existentes na planilha
                df.to_excel(writer, sheet_name=model_name, index=False, header=False, startrow=writer.sheets[model_name].max_row)
            else:
                # Cria uma nova planilha se ela não existir
                df.to_excel(writer, sheet_name=model_name, index=False)

    print(f"Results for model '{model_name}' and k={k} saved to '{file_path}' in sheet '{model_name}'.")

# Continuação do código, incluindo a função main e a chamada principal para main.



def main(config_path, verbose=0):
    """
    Função principal para executar o carregamento do modelo, previsões e salvar os resultados.

    Parameters
    ----------
    config_path : str
        Caminho para o arquivo de configuração.
    verbose : int, optional
        Modo de verbosidade. Se > 0, imprime mensagens de progresso.
    """
    config = load_config(config_path, verbose=verbose)

    if config is None:
        print("Falha no carregamento da configuração.")
        return

    base_model_dir = config['base_model_dir']
    data_base_dir = config['data_base_dir']
    committee_networks = config['committee_networks']
    excel_file = config['excel_file']
    batch_size = config['BATCH']
    input_size = tuple(config['INPUT_SIZE'])

    create_excel_if_not_exists(excel_file)

    for k in range(1, 11):
        if verbose > 0:
            print(f'\nProcessando para k={k}')

        for model_dir in committee_networks:
            model_dir_path = os.path.join(base_model_dir, model_dir)
            model_name = f'{model_dir}_bestLoss_{k}.keras'
            model_path = os.path.join(model_dir_path, model_name)

            if os.path.exists(model_path):
                if verbose > 0:
                    print(f'Carregando modelo e dados para o modelo {model_dir} e k={k}')
                
                model = load_model(model_path, verbose=verbose)
                test_data = load_data_test(data_base_dir, k, BATCH=batch_size, INPUT_SIZE=input_size, verbose=verbose)
                
                if model and test_data:
                    predict_and_save(model, test_data, k, model_dir, file_path=excel_file)
                    
                    if verbose > 0:
                        print(f'Previsões para o modelo {model_dir} salvas no Excel.')
            else:
                if verbose > 0:
                    print(f'O arquivo do modelo {model_name} não existe no diretório {model_dir_path}')

    print("Processamento ensemble concluído.")
    config={'excel_file': excel_file,
            'model_names': committee_networks
    }
    vote.run(config)
    print("Processamento vote concluído.")

if __name__ == "__main__":
    default_yaml = '/media/jczars/4C22F02A22F01B22/Ensembler_pollen_classification/Reports_Ens/config.yaml'
    
    # Argument parser to accept configuration file as parameter
    parser = argparse.ArgumentParser(description='Process model predictions using a configuration file.')
    parser.add_argument('config', nargs='?', default=default_yaml, type=str, help='Path to the YAML configuration file (default: config.yaml)')
    parser.add_argument('--verbose', type=int, default=0, help='Verbosity level (default is 0)')
    
    args = parser.parse_args()
    
    main(args.config, verbose=args.verbose)
    sound_test_finalizado.beep(2)