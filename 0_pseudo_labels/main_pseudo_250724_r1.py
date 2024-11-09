import os
import openpyxl
import tensorflow as tf
import pandas as pd
import sys

# Configuração do caminho para importar módulos personalizados
sys.path.append('/media/jczars/4C22F02A22F01B22/$WinREAgent/Pollen_classification_view/')
print("Caminhos de sistema:", sys.path)

# Importação de módulos e funções
from models import get_data, utils, models_pre, models_train, reports_build
from models import get_calssifica
# Configuração para reduzir a verbosidade do TensorFlow
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
tf.get_logger().setLevel('ERROR')  # Limita as mensagens do TensorFlow para erros somente

# Define o diretório de trabalho
os.chdir('/media/jczars/4C22F02A22F01B22/$WinREAgent/Pollen_classification_view/0_pseudo_labels/')
print("Diretório de trabalho atual:", os.getcwd())

def prepare_data(conf, root_path):
    test_id = conf['id_test']
    model_name = conf['model']
    augmentation = conf['aug']
    base_name = conf['base']
    path_base = conf['path_base']
    labels_path = f"{path_base}/labels"
    #/media/jczars/4C22F02A22F01B22/Ensembler_pollen_classification/BD/BI_5/
    
    # Carrega categorias de rótulos
    categories = sorted(os.listdir(labels_path))

    #Criar as pastas do teste
    utils.create_folders(root_path, flag=0)
    test_folder_name = f"{test_id}_{model_name}_{augmentation}_{base_name}"
    test_folder_path = f"{root_path}/{test_folder_name}"
    utils.create_folders(test_folder_path, flag=0)

    print(f'Save directory for training: {test_folder_path}, ID: {test_folder_name}')

    print('\n[STEP 0].4 - Create the pseudo_csv folder')
    pseudo_csv_path = f"{test_folder_path}/pseudo_csv"
    utils.create_folders(pseudo_csv_path, flag=0)

    print('\n[STEP 0].5 - Create the CSV')
    csv_data_path = f"{path_base}/{base_name}.csv"
    print('CSV data path:', csv_data_path)
    data_csv = utils.create_dataSet(labels_path, csv_data_path, categories)

    num_labels = len(data_csv)
    print('Total number of labeled data:', num_labels)

    print('\n[STEP 0].6 - Split the data')
    path_train, path_val, path_test = get_data.splitData(data_csv, root_path, base_name)

    return {
        'path_train': path_train,
        'path_val': path_val,
        'path_test': path_test,
        'save_dir_train': test_folder_path,
        'pseudo_csv': pseudo_csv_path,
        'size_of_labels': num_labels,
        'categories': categories
    }


def load_data_labels(conf):
    training_data = pd.read_csv(conf['path_train'])
    val_data = pd.read_csv(conf['path_val'])
    img_size = conf['img_size']
    input_size = (img_size, img_size)

    train, val = get_data.load_data_train(training_data, 
                                          val_data, 
                                          conf['aug'], 
                                          input_size)
    return train, val, training_data

def build_train_config(row, res_pre, k):
    save_dir=f"{res_pre['save_dir_train']}/models/"

    return {
        'model': row['model'],
        'id_test': row['id_test'],
        'path_data': row['path_base'],
        'path_test': res_pre['path_test'],
        'batch_size': row['batch_size'],
        'img_size': row['img_size'],
        'num_classes': len(res_pre['categories']),
        'split_valid': row['split_valid'],
        'last_activation': row['last_activation'],
        'save_dir': save_dir,
        'learning_rate': row['learning_rate'],
        'optimizer': row['optimizer'],
        'epochs': row['epochs'],
        'freeze': row['freeze'],
        'k': k
    }


def train_model(config, train_data, val_data):
    model_inst = None
    model_inst = models_pre.hyper_model(config, verbose=1)
    res_train=models_train.run_train(train_data, val_data, model_inst, config)
    

    # Salvar modelo /models
    nm_model=f"{config['id_test']}_{config['model']}"
    nome_model = os.path.join(config['save_dir'], f"{nm_model}_bestLoss_{config['k']}.keras")
    model_inst.save(nome_model)

    return model_inst, res_train

def build_reports_config(time_step, config, res_pre, model_inst, res_train):
    print('\nReports Generation')
    print(f'\n[INFO]--> step 1.4-Evalution tempo {time_step}')
    test_data = pd.read_csv(res_pre['path_test'])
    img_size = config['img_size']
    input_size = (img_size, img_size)
    print('\n[INFO]--> ', input_size)
    test = get_data.load_data_test(test_data, input_size)
    categories=res_pre['categories']
    
    save_dir=f"{res_pre['save_dir_train']}/reports/"
    utils.create_folders(save_dir, 0)
    reports_config={'save_dir':save_dir,
                    'time':time_step,
                    'batch_size': config['batch_size'],
                    'id_test': config['id_test'],
                    'model': config['model']
                    }
    history=res_train['history']
    report_metrics=reports_build.reports_gen(test, model_inst, categories, history, reports_config)

    return report_metrics


def classification(config, res_pre, model, _tempo):
    """
    Classifica imagens não rotuladas e gera pseudo-rótulos.

    Args:
        config (dict): Configurações do modelo e diretórios.
        res_pre (dict): Resultados anteriores, incluindo categorias e caminho de pseudo-rótulos.
        model (torch.nn.Module): Modelo treinado para fazer previsões.
        _tempo (int): Iteração atual do pseudo-labeling.

    Returns:
        pd.DataFrame: DataFrame contendo as previsões de pseudo-rótulos.
    """
    # Definir o caminho para o conjunto de dados não rotulado
    unlabels_path = f"{config['path_base']}/images_unlabels"
    batch_size = config['batch_size']
    categories = res_pre['categories']

    params = {
        'unlabels': unlabels_path,
        'img_size': config['img_size'],
        'batch_size': batch_size,
        'categories': categories
    }

    # Carregar dados não rotulados ou ler CSV com pseudo-rótulos anteriores
    if _tempo == 0:
        print(f"[DEBUG] Carregando imagens não rotuladas do diretório: {unlabels_path}")
        unlabels_generator = get_data.load_unlabels(params)
    else:
        unlabels_csv_path = f"{res_pre['pseudo_csv']}/unlabelSet_T{str(_tempo)}.csv"
        print(f"[DEBUG] Carregando pseudo-rótulos do CSV: {unlabels_csv_path}")
        try:
            df_unlabels = pd.read_csv(unlabels_csv_path)
        except FileNotFoundError:
            print(f"[ERRO] Arquivo não encontrado: {unlabels_csv_path}")
            return None

        if len(df_unlabels) > 0:
            print("[DEBUG] Cabeçalho do DataFrame de pseudo-rótulos:")
            print(df_unlabels.head())
            unlabels_generator = get_data.load_data_test(df_unlabels, input_size=(224, 224))
        else:
            print(f"[AVISO] Nenhum dado encontrado no CSV {unlabels_csv_path}")
            return None

    # Realizar previsões para pseudo-rotulação
    print("[INFO] Realizando pseudo-rotulações na base não rotulada")
    pseudos_df = reports_build.predict_unlabeled_data(
        unlabels_generator, model, batch_size, categories, verbose=2
    )
    print(f"[INFO] Total de pseudo-rótulos gerados: {len(pseudos_df)}")

    return pseudos_df


def selection(data_uns, conf, res_pre, _tempo, training_data):
    """
    Performs selection of pseudo-labels for training if there is unlabeled data available.

    Steps:
    1. Checks if there is any unlabeled data.
    2. Calls `selec` function to select pseudo-labels based on confidence threshold.
    3. Returns a dictionary with paths and sizes of datasets if selection is successful.
    4. Returns 0 if no selection could be made or if there is no unlabeled data.

    Parameters:
        data_uns (DataFrame): Unlabeled data to be processed.
        conf (dict): Configuration dictionary with paths and threshold settings.
        res_pre (dict): Contains the path for saving pseudo-labels.
        _tempo (int): Current time or iteration index.

    Returns:
        dict or int: Returns dictionary with new data paths and dataset sizes if successful; 
                     returns 0 if no selection made or no unlabeled data.
    """
    if not (data_uns.empty):
        print('\n[STEP 2].4- Selection')

        # Perform pseudo-label selection
        #train_data_csv = pd.read_csv(res_pre['pseudo_csv'])
        print('train_data_csv ', training_data)

        res_sel = get_calssifica.selec(
            conf,
            data_uns,
            res_pre['pseudo_csv'], 
            _tempo, 
            training_data,
            conf['limiar']
        )

        if res_sel:
            # Return paths and dataset sizes for further processing
            return {
                '_csv_New_data': res_sel['_csv_New_TrainSet'],
                'path_test': res_pre['path_test'],
                'save_dir_train': conf.get('path_model', ''),  # Assuming model save path in config
                'pseudo_csv': res_pre['pseudo_csv'],
                'ini': res_sel['ini'],
                'select': res_sel['select'],
                'rest': res_sel['rest'],
                'train': res_sel['train'],
                'new_train': res_sel['new_train']
            }
        else:
            # No valid pseudo-labels selected
            return 0
    else:
        # No unlabeled data to process
        return 0

def rel_data(time_step, report_metrics, res_train, res_sel, workbook_path, config_index):
    "--> Salva os dados na planilha"

    print("\n[INFO] Nome da planilha:", workbook_path)
    workbook = openpyxl.load_workbook(workbook_path)
    print("Abas na planilha:", workbook.sheetnames)

    _teste = 'Table'
    print('\n_test ', _teste)
    
    # Check if the sheet already exists
    if _teste in workbook.sheetnames:
        print('Sheet exists.')
        Met_page = workbook[_teste]  # Access the existing sheet
    else:
        print('Creating new sheet.')
        Met_page = workbook.create_sheet(_teste)  # Create a new sheet
        print('[INFO] -rel_data- salvando o cabeçalho do teste')
        cols_exe=['Tempo', 'test_loss', 'test_accuracy',
                        'precisio', 'recall', 'fscore', 'kappa',
                        'str_time', 'end_time', 'delay', 'epoch_finish',
                        'ini', 'select', 'rest', 'train','new_train','id_test']
        Met_page.append(cols_exe)  # Add header row with column names
        
    # Save the workbook after adding or accessing the sheet
    workbook.save(workbook_path)
    print(workbook.sheetnames)  # Visualiza as abas existentes


    data=[str(time_step),
        report_metrics['test_loss'],
        report_metrics['test_accuracy'],
        report_metrics['precision'],
        report_metrics['recall'],
        report_metrics['fscore'],
        report_metrics['kappa'],
        res_train['start_time'],
        res_train['end_time'],
        res_train['duration'],
        res_train['best_epoch'],
        res_sel['ini'],
        res_sel['select'],
        res_sel['rest'],
        res_sel['train'],
        res_sel['new_train'],
        config_index
        ]
    Met_page.append(data)
    workbook.save(workbook_path)


def run(workbook_path, start_index):
    """
    Executa o processo de pseudo-rotulagem e treinamento com base em uma configuração fornecida em uma planilha.
    
    Args:
        workbook_path (str): Caminho para o arquivo de planilha Excel com as configurações.
        start_index (int): Índice de início para processamento das configurações.
    """
    
    print("\n[INFO] Nome da planilha:", workbook_path)
    workbook = openpyxl.load_workbook(workbook_path)
    print("Abas na planilha:", workbook.sheetnames)

    #folder root
    root_path = os.path.dirname(workbook_path)
    

    # Carregar os dados de configuração da aba 'Sheet'
    config_data = pd.read_excel(workbook_path, sheet_name="Sheet")
    
    # Itera sobre cada linha de configuração a partir do índice `start_index`
    for row_idx in range(len(config_data)):
        utils.reset_keras()  # Reseta backend Keras para evitar problemas de memória
        config_index = start_index + row_idx
        config = config_data.loc[config_index]
        print("Configuração atual:", config)

        # Inicializa o controle de tempo e flag para dados não rotulados
        time_step = 0
        has_unlabeled_data = True
        res_pre=prepare_data(config, root_path)
        
        conf_load={'path_train': res_pre['path_train'],
                   'path_val': res_pre['path_val'],
                   'img_size':config['img_size'],
                   'aug':config['aug']}
        
        print(conf_load)
        train, val, training_data=load_data_labels(conf_load)
        

        

        while has_unlabeled_data:
            utils.reset_keras()

            print('\nrun.treinar')
            conf_train=build_train_config(config, res_pre, config_index)
            model_train, res_train=train_model(conf_train, train, val)

            report_metrics=build_reports_config(time_step, config, res_pre, model_train, res_train)

            print('\nrun.classification')
            pseudos_df=classification(config, res_pre, model_train, time_step)            
            
            print('\nrun.selection')
            res_sel=selection(pseudos_df, config, res_pre, time_step, training_data)
            if res_sel is None:
                break

            rel_data(time_step, report_metrics, res_train, res_sel, workbook_path, config_index)
            
            if time_step>0:
                train, val=get_data.reload_data_train(res_sel['_csv_New_TrainSet'])

            time_step += 1


if __name__ == "__main__":
    workbook_path = '/media/jczars/4C22F02A22F01B22/$WinREAgent/Pollen_classification_view/0_pseudo_labels/Reports/config_pseudo_label_pre.xlsx'
    start_index = 0
    run(workbook_path, start_index)
