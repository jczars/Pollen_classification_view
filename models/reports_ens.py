import openpyxl
import pandas as pd
import os, sys
from sklearn.metrics import accuracy_score
os.environ["tf_gpu_allocator"]="cuda_malloc_async"
sys.path.append('/media/jczars/4C22F02A22F01B22/$WinREAgent/Pollen_classification_view/')
print(sys.path)

from models import reports_build as reports
from models import  utils

def reports_gen(y_true, y_pred, categories, df_correct, nm_model, k, save_dir=None):
    
    # Confusion matrix
    matrix_fig, mat = reports.plot_confusion_matrixV4(y_true, y_pred, categories, normalize=None)
    df_mat = pd.DataFrame(mat, index=categories, columns=categories)
    
    # Boxplot, classification report and training metrics
    boxplot_fig = reports.plot_confidence_boxplot(df_correct)
    class_report = reports.generate_classification_report(y_true, y_pred, categories)
        
    # Save files if directory is specified
    if save_dir:
        folder_name=f"{nm_model}_reports/"
        save_dir = os.path.join(save_dir, folder_name)
        print(save_dir)
        utils.create_folders(save_dir, flag=0)
        print("save graph")
        df_mat.to_csv(f'{save_dir}/Test_{nm_model}_mat_conf_k{k}.csv')
        df_correct.to_csv(f'{save_dir}/Test_{nm_model}_df_correct_k{k}.csv')
        
        class_report.to_csv(f'{save_dir}/Test_{nm_model}_Class_reports_k{k}.csv')
                
        matrix_fig.savefig(f'{save_dir}/Test_{nm_model}_mat_conf_k{k}.jpg')
        boxplot_fig.savefig(f'{save_dir}/Test_{nm_model}_boxplot_k{k}.jpg')
    
    # Calculate metrics
    me = reports.calculate_metrics(y_true, y_pred)
    accuracy = accuracy_score(y_true, y_pred)

    # Return evaluation metrics
    me = {
        'test_accuracy': accuracy,
        'precision': me['precision'],
        'recall': me['recall'],
        'fscore': me['fscore'],
        'kappa': me['kappa'],
    }
    
    return me

def run(file_path):
    # Carrega o workbook uma vez para verificar as abas existentes
    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames

    for i in range(1, 11):  # Tentativa de processar até k10
        sheet = f'k{i}'
        
        # Verifica se a aba existe
        if sheet not in sheet_names:
            print(f"Aba '{sheet}' não encontrada. Parando o processamento.")
            break  # Encerra o loop se uma aba não existir
        
        print('Processando', sheet)
        df = pd.read_excel(file_path, sheet_name=sheet)

        # Exibe as primeiras linhas do DataFrame
        print(df.head()) 
        y_true = df['y_true']
        y_pred = df['elected']   
        categories = df['y_true'].unique()  
        df_correct = df[df['status'] == 'Correct'] 
        # Renomeia a coluna 'probability' para 'confidence'
        df_correct = df_correct.rename(columns={'probability': 'confidence', 'y_true': 'true_label'})
        nm_model = file_path.split('/')[-1].split('.')[0]
        
        save_dir = os.path.dirname(file_path)   
        print(save_dir)
        me = reports_gen(y_true, y_pred, categories, df_correct, nm_model, i, save_dir)  
        
        # Adiciona as métricas na aba 'metrics'
        sheet_name = 'metrics'
        
        # Cria a aba apenas na primeira execução
        if sheet_name not in workbook.sheetnames:
            Met_page = workbook.create_sheet(sheet_name)
            Met_page.append(['k', 'test_accuracy', 'precision', 'recall', 'fscore', 'kappa'])  # Cabeçalho
        
        # Adiciona as métricas como uma nova linha
        Met_page = workbook[sheet_name]
        Met_page.append([i, me['test_accuracy'], me['precision'], me['recall'], me['fscore'], me['kappa']])
        
        # Salva o workbook atualizado
        workbook.save(file_path)

        print('Processo de relatórios completo para', sheet)

if __name__ == "__main__":
    # Caminho para o arquivo Excel
    file_path = '3_ensemble/Reports/POLAR_ens_111124.xlsx'

    run(file_path)
    
